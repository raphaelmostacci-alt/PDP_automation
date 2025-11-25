"""
Module de génération de rapports Excel
Crée un fichier Excel avec le statut de conformité de tous les documents
"""

from pathlib import Path
from datetime import datetime
from typing import List, Dict
import logging

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None
    logging.warning("⚠️ openpyxl non installé. Génération Excel non disponible.")

from config import (
    OUTPUT_DIR,
    get_excel_filename,
    EXCEL_COLUMNS,
    EXCEL_SHEET_NAME,
    STATUS_CONFORME,
    STATUS_NON_CONFORME,
    STATUS_ERREUR,
    STATUS_A_VERIFIER
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelGenerator:
    """Générateur de rapports Excel pour les PDP"""
    
    def __init__(self, output_dir: Path = OUTPUT_DIR):
        """
        Initialise le générateur Excel
        
        Args:
            output_dir: Dossier de sortie pour les rapports
        """
        if not Workbook:
            raise ImportError("openpyxl n'est pas installé. Lancez: pip install openpyxl")
        
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        self.workbook = None
        self.worksheet = None
    
    def create_report(self, results: List[Dict], filename: str = None) -> Path:
        """
        Crée un rapport Excel à partir des résultats de validation
        
        Args:
            results: Liste des résultats de validation
            filename: Nom du fichier (optionnel, auto-généré si None)
        
        Returns:
            Chemin du fichier généré
        """
        logger.info("📊 Création du rapport Excel...")
        
        # Créer le workbook
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = EXCEL_SHEET_NAME
        
        # Ajouter l'en-tête
        self._add_header()
        
        # Ajouter les données
        self._add_data(results)
        
        # Appliquer la mise en forme
        self._apply_formatting()
        
        # Ajouter les statistiques
        self._add_statistics(results)
        
        # Sauvegarder
        if not filename:
            filename = get_excel_filename()
        
        output_path = self.output_dir / filename
        self.workbook.save(output_path)
        
        logger.info(f"✅ Rapport sauvegardé: {output_path}")
        return output_path
    
    def _add_header(self):
        """Ajoute l'en-tête du tableau"""
        for col_num, column_title in enumerate(EXCEL_COLUMNS, 1):
            cell = self.worksheet.cell(row=1, column=col_num)
            cell.value = column_title
            
            # Style de l'en-tête
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def _add_data(self, results: List[Dict]):
        """
        Ajoute les données au tableau
        
        Args:
            results: Liste des résultats
        """
        for row_num, result in enumerate(results, 2):
            # Extraire les données
            entreprise = result.get('entreprise', 'Non spécifié')
            nom = result.get('nom', 'Non trouvé')
            prenom = result.get('prenom', 'Non trouvé')
            doc_type = result.get('doc_type', 'UNKNOWN')
            file_name = result.get('file_name', '')
            date_validite = result.get('date_validite', 'N/A')
            statut = result.get('statut', STATUS_ERREUR)
            commentaire = result.get('commentaire', '')
            
            # Remplir la ligne
            row_data = [
                entreprise,
                nom,
                prenom,
                doc_type,
                file_name,
                date_validite,
                statut,
                commentaire
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = self.worksheet.cell(row=row_num, column=col_num)
                cell.value = value
                
                # Colorer la cellule statut selon le résultat
                if col_num == 7:  # Colonne Statut
                    self._apply_status_color(cell, statut)
    
    def _apply_status_color(self, cell, statut: str):
        """
        Applique une couleur selon le statut
        
        Args:
            cell: Cellule Excel
            statut: Statut de conformité
        """
        colors = {
            STATUS_CONFORME: "C6EFCE",      # Vert clair
            STATUS_NON_CONFORME: "FFC7CE",  # Rouge clair
            STATUS_ERREUR: "FFEB9C",        # Jaune clair
            STATUS_A_VERIFIER: "E7E6E6"     # Gris clair
        }
        
        font_colors = {
            STATUS_CONFORME: "006100",      # Vert foncé
            STATUS_NON_CONFORME: "9C0006",  # Rouge foncé
            STATUS_ERREUR: "9C6500",        # Jaune foncé
            STATUS_A_VERIFIER: "3F3F3F"     # Gris foncé
        }
        
        fill_color = colors.get(statut, "FFFFFF")
        font_color = font_colors.get(statut, "000000")
        
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        cell.font = Font(color=font_color, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def _apply_formatting(self):
        """Applique la mise en forme générale"""
        # Ajuster la largeur des colonnes
        column_widths = {
            'A': 20,  # Entreprise
            'B': 15,  # Nom
            'C': 15,  # Prénom
            'D': 20,  # Type Document
            'E': 30,  # Fichier
            'F': 15,  # Date Validité
            'G': 18,  # Statut
            'H': 40   # Commentaire
        }
        
        for col, width in column_widths.items():
            self.worksheet.column_dimensions[col].width = width
        
        # Figer la première ligne
        self.worksheet.freeze_panes = 'A2'
        
        # Ajouter des bordures
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in self.worksheet.iter_rows(min_row=1, max_row=self.worksheet.max_row, 
                                            min_col=1, max_col=len(EXCEL_COLUMNS)):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:  # Pas l'en-tête
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
    
    def _add_statistics(self, results: List[Dict]):
        """
        Ajoute un tableau de statistiques en bas du rapport
        
        Args:
            results: Liste des résultats
        """
        # Calculer les statistiques
        total = len(results)
        conformes = sum(1 for r in results if r.get('statut') == STATUS_CONFORME)
        non_conformes = sum(1 for r in results if r.get('statut') == STATUS_NON_CONFORME)
        erreurs = sum(1 for r in results if r.get('statut') == STATUS_ERREUR)
        a_verifier = sum(1 for r in results if r.get('statut') == STATUS_A_VERIFIER)
        
        taux_conformite = (conformes / total * 100) if total > 0 else 0
        
        # Position du tableau de stats (3 lignes après les données)
        stats_start_row = self.worksheet.max_row + 3
        
        # Titre
        cell = self.worksheet.cell(row=stats_start_row, column=1)
        cell.value = "STATISTIQUES"
        cell.font = Font(bold=True, size=12)
        
        # Données statistiques
        stats_data = [
            ("Total documents analysés:", total),
            ("✅ Conformes:", conformes),
            ("❌ Non conformes:", non_conformes),
            ("⚠️  Erreurs:", erreurs),
            ("🔍 À vérifier:", a_verifier),
            ("📈 Taux de conformité:", f"{taux_conformite:.1f}%")
        ]
        
        for i, (label, value) in enumerate(stats_data, stats_start_row + 1):
            self.worksheet.cell(row=i, column=1, value=label).font = Font(bold=True)
            self.worksheet.cell(row=i, column=2, value=value)
        
        # Ajouter la date de génération
        date_row = stats_start_row + len(stats_data) + 2
        cell = self.worksheet.cell(row=date_row, column=1)
        cell.value = f"Rapport généré le: {datetime.now().strftime('%d/%m/%Y à %H:%M:%S')}"
        cell.font = Font(italic=True, size=9)


def main():
    """Fonction de test du générateur Excel"""
    # Données de test
    test_results = [
        {
            'entreprise': 'Entreprise A',
            'nom': 'DUPONT',
            'prenom': 'Jean',
            'doc_type': 'CNI',
            'file_name': 'CNI_DUPONT.pdf',
            'date_validite': '31/12/2027',
            'statut': STATUS_CONFORME,
            'commentaire': 'CNI valide jusqu\'au 31/12/2027'
        },
        {
            'entreprise': 'Entreprise A',
            'nom': 'MARTIN',
            'prenom': 'Marie',
            'doc_type': 'CNI',
            'file_name': 'CNI_MARTIN.pdf',
            'date_validite': '15/06/2020',
            'statut': STATUS_NON_CONFORME,
            'commentaire': 'CNI expirée le 15/06/2020'
        },
        {
            'entreprise': 'Entreprise B',
            'nom': 'DURAND',
            'prenom': 'Paul',
            'doc_type': 'HABILITATION_ELEC',
            'file_name': 'HAB_DURAND.pdf',
            'date_validite': '15/01/2026',
            'statut': STATUS_CONFORME,
            'commentaire': 'Habilitation valide jusqu\'au 15/01/2026'
        },
        {
            'entreprise': 'Entreprise C',
            'nom': 'Non trouvé',
            'prenom': 'Non trouvé',
            'doc_type': 'FDS',
            'file_name': 'FDS_Acetone.pdf',
            'date_validite': 'N/A',
            'statut': STATUS_ERREUR,
            'commentaire': 'Impossible d\'extraire les données'
        }
    ]
    
    try:
        generator = ExcelGenerator()
        output_file = generator.create_report(test_results, "Test_Rapport_PDP.xlsx")
        print(f"\n✅ Rapport de test créé: {output_file}")
        print("📂 Ouvrez-le pour vérifier la mise en forme!")
        
    except Exception as e:
        logger.error(f"❌ Erreur lors de la génération: {e}")
        raise


if __name__ == "__main__":
    main()
